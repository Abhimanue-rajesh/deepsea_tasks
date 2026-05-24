import os

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.files import File
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from credentials.models import Credential


class Command(BaseCommand):
    help = "Import credentials from Excel file"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)

    def handle(self, *args, **options):
        excel_path = options["excel_path"]

        self.stdout.write(self.style.WARNING("Starting credential import..."))
        self.stdout.write(f"Excel file: {excel_path}")

        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR("Excel file not found."))
            return

        wb = load_workbook(excel_path)
        sheet = wb.active

        headers = [
            str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]
        ]
        self.stdout.write(f"Detected headers: {headers}")

        imported_count = 0
        skipped_count = 0
        error_count = 0

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if all(value is None for value in row):
                continue
            data = {
                key: str(value).strip() if value is not None else None
                for key, value in zip(headers, row)
            }

            name = data.get("name")
            user_id = data.get("user_id")
            email = data.get("email")
            password = data.get("password")
            url = data.get("url")
            pem_file_name = data.get("pem_file")
            notes = data.get("notes")

            if not name:
                skipped_count += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"Row {row_number}: Skipped because name is empty"
                    )
                )
                continue

            credential, created = Credential.objects.update_or_create(
                name=name,
                defaults={
                    "user_id": user_id,
                    "email": email,
                    "password": password,
                    "url": url,
                    "notes": notes,
                },
            )

            try:
                if pem_file_name:
                    pem_path = os.path.join(
                        settings.MEDIA_ROOT,
                        "import_pem_files",
                        str(pem_file_name),
                    )

                    if os.path.exists(pem_path):
                        with open(pem_path, "rb") as f:
                            credential.pem_file.save(
                                str(pem_file_name),
                                File(f),
                                save=False,
                            )
                    else:
                        self.stdout.write(
                            self.style.WARNING(
                                f"Row {row_number}: PEM file not found - {pem_path}"
                            )
                        )

                credential.full_clean()
                credential.save()

                imported_count += 1
                self.stdout.write(
                    self.style.SUCCESS(
                        f"Row {row_number}: Imported - {credential.name}"
                    )
                )

            except ValidationError as e:
                error_count += 1
                self.stdout.write(
                    self.style.ERROR(
                        f"Row {row_number}: Validation error - {e.message_dict}"
                    )
                )

            except Exception as e:
                error_count += 1
                self.stdout.write(
                    self.style.ERROR(f"Row {row_number}: Error - {str(e)}")
                )

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Import completed."))
        self.stdout.write(f"Imported: {imported_count}")
        self.stdout.write(f"Skipped: {skipped_count}")
        self.stdout.write(f"Errors: {error_count}")
